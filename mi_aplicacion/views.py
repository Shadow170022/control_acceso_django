from django.shortcuts import render
from django.db.models import Q, Min, Max
from .models import Empleado
from datetime import date, datetime, timedelta
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import permission_required
from django.shortcuts import redirect
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', '/rh/empleados_por_fecha/')  # permite la captura el parámetro 'next'
            return redirect(next_url)
        else:
            messages.error(request, 'Credenciales inválidas')
    return render(request, 'rh_login.html')

def logout_view(request):
    logout(request)
    return redirect('rh_login')

@login_required(login_url='/rh/login/')
@permission_required('mi_aplicacion.is_rh', login_url='/rh/login/', raise_exception=True)
def empleados_por_fecha(request):
    """View to filter employees by name and/or date range and group their working hours."""
    nombre = request.GET.get('nombre', '').strip()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    agrupar_total = request.GET.get('agrupar_total') == 'on'

    # Convert date strings to date objects
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
    except ValueError:
        start_date, end_date = None, None

    empleados = []
    auth_data = []
    total_horas_trabajadas = None

    if (start_date and end_date) or nombre:
        empleados_queryset = Empleado.objects.all()

        if start_date and end_date:
            empleados_queryset = empleados_queryset.filter(auth_date__range=(start_date, end_date))

        if nombre:
            empleados_queryset = empleados_queryset.filter(Q(user_name__icontains=nombre))

        if agrupar_total:
            auth_data = list(empleados_queryset.values('user_name', 'id_empleado', 'auth_date', 'auth_time'))

            auth_totals = {}
            for empleado in auth_data:
                user_name = empleado['user_name']
                id_empleado = empleado['id_empleado']
                auth_date = empleado['auth_date']
                auth_time = empleado['auth_time']

                if (user_name, id_empleado) not in auth_totals:
                    auth_totals[(user_name, id_empleado)] = {}

                if auth_date not in auth_totals[(user_name, id_empleado)]:
                    auth_totals[(user_name, id_empleado)][auth_date] = []

                if auth_time:
                    auth_totals[(user_name, id_empleado)][auth_date].append(auth_time)

            empleados = [
                {
                    'user_name': key[0],
                    'id_empleado': key[1],
                    'auth_dates': value,
                    'total_horas_trabajadas': timedelta()
                }
                for key, value in auth_totals.items()
            ]

            for empleado in empleados:
                for auth_date, auth_times in empleado['auth_dates'].items():
                    auth_times.sort()
                    first_checkin = auth_times[0]
                    last_checkout = auth_times[-1]

                    first_checkin_dt = datetime.combine(datetime.min, first_checkin)
                    last_checkout_dt = datetime.combine(datetime.min, last_checkout)
                    horas_trabajadas = last_checkout_dt - first_checkin_dt
                    empleado['total_horas_trabajadas'] += horas_trabajadas

            for empleado in empleados:
                total_seconds = int(empleado['total_horas_trabajadas'].total_seconds())
                horas_totales = total_seconds // 3600
                minutos_totales = (total_seconds % 3600) // 60
                segundos_totales = total_seconds % 60
                empleado['total_horas_trabajadas'] = f"{horas_totales:02}:{minutos_totales:02}:{segundos_totales:02}"

        else:
            empleados = empleados_queryset.values('user_name', 'id_empleado', 'auth_date').annotate(
                first_checkin=Min('auth_time'),
                last_checkout=Max('auth_time')
            ).order_by('auth_date', 'id_empleado')

            for empleado in empleados:
                checkin = empleado['first_checkin']
                checkout = empleado['last_checkout']

                if checkin and checkout:
                    checkin_datetime = datetime.combine(empleado['auth_date'], checkin)
                    checkout_datetime = datetime.combine(empleado['auth_date'], checkout)
                    horas_trabajadas = (checkout_datetime - checkin_datetime).total_seconds() / 3600
                    empleado['horas_trabajadas'] = round(horas_trabajadas, 2)
                else:
                    empleado['horas_trabajadas'] = "N/A"

    formatted_start_date = start_date.strftime('%Y-%m-%d') if start_date else ''
    formatted_end_date = end_date.strftime('%Y-%m-%d') if end_date else ''

    return render(request, 'empleados_por_fecha.html', {
        'empleados': empleados,
        'start_date': formatted_start_date,
        'end_date': formatted_end_date,
        'nombre': nombre,
        'agrupar_total': agrupar_total,
        'auth_data': auth_data if agrupar_total else None,
        'total_horas_trabajadas': total_horas_trabajadas if agrupar_total else None
    })

@login_required(login_url='/rh/login/')
@permission_required('mi_aplicacion.is_rh', login_url='/rh/login/', raise_exception=True)
def employees_view(request):
    """View to filter employees by name and/or date range and group their working hours."""
    nombre = request.GET.get('nombre', '').strip()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    agrupar_total = request.GET.get('agrupar_total') == 'on'

    # Convert date strings to date objects
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
    except ValueError:
        start_date, end_date = None, None

    empleados = []
    auth_data = []
    total_horas_trabajadas = None

    if (start_date and end_date) or nombre:
        empleados_queryset = Empleado.objects.all()

        if start_date and end_date:
            empleados_queryset = empleados_queryset.filter(auth_date__range=(start_date, end_date))

        if nombre:
            empleados_queryset = empleados_queryset.filter(Q(user_name__icontains=nombre))

        if agrupar_total:
            auth_data = list(empleados_queryset.values('user_name', 'id_empleado', 'auth_date', 'auth_time'))

            auth_totals = {}
            for empleado in auth_data:
                user_name = empleado['user_name']
                id_empleado = empleado['id_empleado']
                auth_date = empleado['auth_date']
                auth_time = empleado['auth_time']

                if (user_name, id_empleado) not in auth_totals:
                    auth_totals[(user_name, id_empleado)] = {}

                if auth_date not in auth_totals[(user_name, id_empleado)]:
                    auth_totals[(user_name, id_empleado)][auth_date] = []

                if auth_time:
                    auth_totals[(user_name, id_empleado)][auth_date].append(auth_time)

            empleados = [
                {
                    'user_name': key[0],
                    'id_empleado': key[1],
                    'auth_dates': value,
                    'total_horas_trabajadas': timedelta()
                }
                for key, value in auth_totals.items()
            ]

            for empleado in empleados:
                for auth_date, auth_times in empleado['auth_dates'].items():
                    auth_times.sort()
                    first_checkin = auth_times[0]
                    last_checkout = auth_times[-1]

                    first_checkin_dt = datetime.combine(datetime.min, first_checkin)
                    last_checkout_dt = datetime.combine(datetime.min, last_checkout)
                    horas_trabajadas = last_checkout_dt - first_checkin_dt
                    empleado['total_horas_trabajadas'] += horas_trabajadas

            for empleado in empleados:
                total_seconds = int(empleado['total_horas_trabajadas'].total_seconds())
                horas_totales = total_seconds // 3600
                minutos_totales = (total_seconds % 3600) // 60
                segundos_totales = total_seconds % 60
                empleado['total_horas_trabajadas'] = f"{horas_totales:02}:{minutos_totales:02}:{segundos_totales:02}"

        else:
            empleados = empleados_queryset.values('user_name', 'id_empleado', 'auth_date').annotate(
                first_checkin=Min('auth_time'),
                last_checkout=Max('auth_time')
            ).order_by('auth_date', 'id_empleado')

            for empleado in empleados:
                checkin = empleado['first_checkin']
                checkout = empleado['last_checkout']

                if checkin and checkout:
                    checkin_datetime = datetime.combine(empleado['auth_date'], checkin)
                    checkout_datetime = datetime.combine(empleado['auth_date'], checkout)
                    horas_trabajadas = (checkout_datetime - checkin_datetime).total_seconds() / 3600
                    empleado['horas_trabajadas'] = round(horas_trabajadas, 2)
                else:
                    empleado['horas_trabajadas'] = "N/A"

    formatted_start_date = start_date.strftime('%Y-%m-%d') if start_date else ''
    formatted_end_date = end_date.strftime('%Y-%m-%d') if end_date else ''

    # Convertimos el QuerySet a una lista serializable
    empleados_data = list(empleados)

    return render(request, 'employees_view.html', {
        'empleados': empleados_data,
        'start_date': formatted_start_date,
        'end_date': formatted_end_date,
        'nombre': nombre,
        'agrupar_total': agrupar_total,
        'auth_data': auth_data if agrupar_total else None,
        'total_horas_trabajadas': total_horas_trabajadas if agrupar_total else None
    })


@login_required(login_url='/rh/login/')
@permission_required('mi_aplicacion.is_rh', login_url='/rh/login/', raise_exception=True)
def export_employees_report(request):
    nombre = request.GET.get('nombre', '').strip()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    agrupar_total = request.GET.get('agrupar_total') == 'on'

    # Convert date strings to date objects
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
    except ValueError:
        start_date, end_date = None, None

    empleados = []
    auth_data = []
    total_horas_trabajadas = None

    if (start_date and end_date) or nombre:
        empleados_queryset = Empleado.objects.all()

        if start_date and end_date:
            empleados_queryset = empleados_queryset.filter(auth_date__range=(start_date, end_date))

        if nombre:
            empleados_queryset = empleados_queryset.filter(Q(user_name__icontains=nombre))

        if agrupar_total:
            # Nuevo agrupamiento por fecha y empleado
            auth_data = list(empleados_queryset.values(
                'user_name', 
                'id_empleado', 
                'auth_date', 
                'auth_time'
            ).order_by('auth_date'))

            auth_totals = {}
            for empleado in auth_data:
                key = (
                    empleado['auth_date'],
                    empleado['user_name'],
                    empleado['id_empleado']
                )
                
                if key not in auth_totals:
                    auth_totals[key] = []
                
                if empleado['auth_time']:
                    auth_totals[key].append(empleado['auth_time'])

            empleados = []
            for key, auth_times in auth_totals.items():
                auth_date, user_name, id_empleado = key
                auth_times.sort()
                
                first_checkin = auth_times[0]
                last_checkout = auth_times[-1]
                
                # Cálculo de horas trabajadas
                checkin_dt = datetime.combine(auth_date, first_checkin)
                checkout_dt = datetime.combine(auth_date, last_checkout)
                horas_trabajadas = checkout_dt - checkin_dt
                
                total_seconds = int(horas_trabajadas.total_seconds())
                horas = total_seconds // 3600
                minutos = (total_seconds % 3600) // 60
                segundos = total_seconds % 60
                
                empleados.append({
                    'auth_date': auth_date,
                    'user_name': user_name,
                    'id_empleado': id_empleado,
                    'first_checkin': first_checkin.strftime('%H:%M:%S'),
                    'last_checkout': last_checkout.strftime('%H:%M:%S'),
                    'horas_trabajadas': f"{horas:02}:{minutos:02}:{segundos:02}"
                })

        else:
            # Consulta original con agrupamiento por fecha
            empleados = empleados_queryset.values(
                'auth_date',
                'user_name', 
                'id_empleado'
            ).annotate(
                first_checkin=Min('auth_time'),
                last_checkout=Max('auth_time')
            ).order_by('auth_date', 'id_empleado')

            for empleado in empleados:
                checkin = empleado['first_checkin']
                checkout = empleado['last_checkout']
                fecha = empleado['auth_date']
                
                if checkin and checkout:
                    checkin_dt = datetime.combine(fecha, checkin)
                    checkout_dt = datetime.combine(fecha, checkout)
                    horas_trabajadas = (checkout_dt - checkin_dt).total_seconds()
                    
                    horas = int(horas_trabajadas) // 3600
                    minutos = (int(horas_trabajadas) % 3600) // 60
                    segundos = int(horas_trabajadas) % 60
                    
                    empleado['horas_trabajadas'] = f"{horas:02}:{minutos:02}:{segundos:02}"
                else:
                    empleado['horas_trabajadas'] = "N/A"

    # Ordenar por fecha antes de exportar
    empleados_data = sorted(
        empleados, 
        key=lambda x: x['auth_date'] if x.get('auth_date') else date.min
    )

    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Report"

    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007bff", fill_type="solid")
    alignment = Alignment(vertical='center', wrap_text=True)

    # Nuevos encabezados con columna de fecha
    headers = [
        ('Fecha', 15),
        ('ID Empleado', 15),
        ('Nombre', 25),
        ('Primer Check-In', 20),
        ('Último Check-Out', 20),
        ('Horas Trabajadas', 15)
    ]

    # Escribir encabezados
    for col_num, (header, width) in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        ws.column_dimensions[col_letter].width = width

    # Escribir datos
    for row_num, empleado in enumerate(empleados_data, 2):
        ws[f'A{row_num}'] = empleado.get('auth_date')
        ws[f'A{row_num}'].number_format = 'YYYY-MM-DD'
        ws[f'B{row_num}'] = empleado.get('id_empleado')
        ws[f'C{row_num}'] = empleado.get('user_name')
        ws[f'D{row_num}'] = empleado.get('first_checkin', 'N/A')
        ws[f'E{row_num}'] = empleado.get('last_checkout', 'N/A')
        ws[f'F{row_num}'] = empleado.get('horas_trabajadas', 'N/A')
        
        # Determinar qué campo de horas usar
        horas = empleado.get('total_horas_trabajadas') if agrupar_total \
               else empleado.get('horas_trabajadas', 'N/A')
        ws[f'E{row_num}'] = horas

    # Autoajustar filas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="reporte_empleados_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'},
    )
    
    wb.save(response)
    return response

def lista_empleados(request):
    """View to list employees and calculate their working hours for the current day."""
    today = date.today()

    # Group by employee and get the first and last authentication record of the day
    empleados = Empleado.objects.filter(auth_date=today).values('user_name', 'id_empleado').annotate(
        first_checkin=Min('auth_time'),
        last_checkout=Max('auth_time')
    )

    # Calculate working hours if both check-in and check-out times exist
    for empleado in empleados:
        checkin = empleado['first_checkin']
        checkout = empleado['last_checkout']

        if checkin and checkout:
            checkin_datetime = datetime.combine(today, checkin)
            checkout_datetime = datetime.combine(today, checkout)
            horas_trabajadas = (checkout_datetime - checkin_datetime).total_seconds() / 3600  # Convert to hours
            empleado['horas_trabajadas'] = round(horas_trabajadas, 2)
        else:
            empleado['horas_trabajadas'] = "N/A"

    return render(request, 'empleados.html', {'empleados': empleados, 'today': today})
